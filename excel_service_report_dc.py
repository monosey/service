import openpyxl
import sql_service_report_dc




def uto_5_excel(station_name, type_traffic, from_month, to_month, to_date, from_date):
    uto5 = sql_service_report_dc.uto_5_sql(station_name, type_traffic, from_month, to_month, to_date, from_date)
    print(uto5)
    
    work_book = openpyxl.Workbook()
    
    sheet = work_book.get_sheet_by_name('Sheet')
    uto5_row = 2
    
    sheet.cell(row = 1, column = 1).value = 'Станция'
    sheet.cell(row = 1, column = 2).value = 'Вид движения'
    sheet.cell(row = 1, column = 3).value = 'Фамилия машиниста'
    
    for i in uto5:
        
        sheet.cell(row = uto5_row, column = 1).value = i['station']
        sheet.cell(row = uto5_row, column = 2).value = i['type_traffic']
        sheet.cell(row = uto5_row, column = 3).value = i['fio']
        
        
        
        uto5_row += 1
    work_book.close()
    work_book.save(filename = 'file/import_excel_uto5.xlsx')
    